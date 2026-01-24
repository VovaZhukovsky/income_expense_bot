import common
import matches
import os
from openpyxl import load_workbook
from telegram import Update,InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext, CallbackQueryHandler
from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
from datetime import datetime, date, timedelta

# to-do объединить с income
async def expense(update: Update, context: CallbackContext):
    context.user_data["mode"] = "expense"
    keyboard = [
        [InlineKeyboardButton("✏️ Изменить категорию", callback_data=f"get_expense_categories")],
        [InlineKeyboardButton("📅 Изменить дату", callback_data=f"set_date")],
        [InlineKeyboardButton("➕ Добавить расход", callback_data=f"ask_for_expense_increment")],
        [InlineKeyboardButton("➖ Уменьшить расход", callback_data=f"ask_for_expense_decrement")],
        [InlineKeyboardButton("🔍 Посмотреть расход", callback_data=f"get_expense_view")]
    ]
    if not os.path.isfile(common.local_xlsx_path):
        client = common.get_ya_client(context._user_id)
        with client:
            client.download(common.ya_xlsx_path, common.local_xlsx_path)
    workbook = load_workbook(common.local_xlsx_path)
    sheet = workbook[common.year]
    context.user_data['sheet'] = sheet
    context.user_data['workbook'] = workbook

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        if context.user_data.get("selected_expense_category") is None:
            context.user_data['selected_expense_category'] = common.DEFAULT_EXPENSE_INFO
        if context.user_data.get("selected_date") is None:
            context.user_data['selected_date'] = date.today()
    else:
        context.user_data['selected_expense_category'] = common.DEFAULT_EXPENSE_INFO
        context.user_data['selected_date'] = date.today()

    if update.callback_query:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            "Расход\n"
            f"Категория: {context.user_data['selected_expense_category']['name']}\n"
            f"Дата: {context.user_data['selected_date']}\n\n",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            "Расход\n"
            f"Категория: {context.user_data['selected_expense_category']['name']}\n"
            f"Дата: {context.user_data['selected_date']}\n\n",
            reply_markup=reply_markup
        )

async def get_expense_categories(update: Update, context: CallbackContext):
    sheet = context.user_data['sheet']
    category_list = []
    index = 2
    for col in sheet.iter_cols(min_col=2,min_row=2, max_col=19, values_only=True):
        category_list.append({"id": index, "name": col[0], "month": matches.get_month(index)})
        index += 1

    context.user_data['category_list'] = category_list
    
    keyboard = []
   
    for category in category_list:
        keyboard.append([
            InlineKeyboardButton(
                f"{category['name']}",
                callback_data=f"select_expense_category_{category['id']}"
            )
        ])
    keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_expense")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            "Выберите категорию:",
            reply_markup=reply_markup
        )

async def select_expense_category_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    await query.answer()

    # Извлекаем ID категории из callback_data
    category_id = int(query.data.replace("select_expense_category_", ""))

    # Находим категорию в списке
    category_list = context.user_data.get('category_list', [])
    selected_category = next((c for c in category_list if c['id'] == category_id), None)

    if not selected_category:
        await query.edit_message_text("Категория не найдена!")
        return

    # Сохраняем выбранную категорию в контексте
    context.user_data['selected_expense_category'] = selected_category
    await expense(update, context)

async def process_calendar_callback(update: Update, context: CallbackContext):
    """Обработчик выбора даты из календаря"""
    query = update.callback_query
    await query.answer()
    
    calendar = context.user_data.get('calendar')
    
    if not calendar:
        calendar = DetailedTelegramCalendar()
    
    # Обрабатываем выбор в календаре
    result, key, step = calendar.process(query.data)
    
    if not result and key:
        # Обновляем календарь для следующего шага
        await query.edit_message_text(
            f"📅 Выберите дату ({LSTEP[step]}):",
            reply_markup=key
        )
    elif result:
        # Дата выбрана
        selected_date = result
        context.user_data['selected_date'] = selected_date

        await expense(update, context)

async def ask_for_expense(update: Update, context: CallbackContext):
    """Запрашиваем расход"""


    query = update.callback_query
    await query.answer()

    operator = query.data.replace("ask_for_expense_", "")
    context.user_data['operator'] = operator

    keyboard = [
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_expense")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        f"Категория: {context.user_data['selected_expense_category']['name']}\n"
        f"Дата: {context.user_data['selected_date']}\n\n"
        "Введите сумму расхода:",
        reply_markup=reply_markup)

async def get_expense_view(update: Update, context: CallbackContext):
    """Запрашиваем просмотр расхода"""
    query = update.callback_query
    await query.answer()

    keyboard = [
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_expense")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        sheet = context.user_data.get('sheet')
        selected_category = context.user_data.get('selected_expense_category')
        # Запись данных
        month = context.user_data['selected_date'].strftime("%B")
        row = matches.get_month(month)
        col = matches.get_day_number(selected_category['id'])
        current_value = sheet[f'{col}{row}'].value
        if current_value is None:
            current_value = 0
        
        await query.edit_message_text(
                f"✅ Успешно!\n\n"
                f"Категория: {context.user_data['selected_expense_category']['name']}\n"
                f"📅 Дата: {context.user_data['selected_date']}\n"
                f"💰 Текущий расход: {current_value}\n\n"
                "Что дальше?",
                reply_markup=reply_markup
            )

    except ValueError:
        await update.message.reply_text("При получении инфы о расходе произошла ошибка.")

async def process_expense_input(update: Update, context: CallbackContext):
    """Обрабатывает введенное количество часов"""

    try:
        # Парсим введенное значение
        expense_text = update.message.text.strip()
        expense = float(update.message.text.strip().replace(',', '.'))
        
        # Проверяем валидность
        if expense <= 0:
            await update.message.reply_text("Пожалуйста, введите положительное число:")
            return
        
        success = await backend_add_expense_to_timesheet(context=context, expense=expense)
        
        if success:
            keyboard = [
                [InlineKeyboardButton("➕ Добавить расход", callback_data=f"ask_for_expense_increment")],
                [InlineKeyboardButton("➖ Уменьшить расход", callback_data=f"ask_for_expense_decrement")],
                [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_expense")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"✅ Успешно добавлено!\n\n"
                f"Категория: {context.user_data['selected_expense_category']['name']}\n"
                f"📅 Дата: {context.user_data['selected_date']}\n"
                f"💰 Расход: {expense}\n\n"
                "Что дальше?",
                reply_markup=reply_markup
            )
            client = common.get_ya_client(context._user_id)
            with client:
                client.upload(common.local_xlsx_path, common.ya_xlsx_path,overwrite=True)
            os.remove(common.local_xlsx_path)
        else:
            common.logger.error(f"Ошибка при добавлении расхода: {expense}")
            await update.message.reply_text("❌ Ошибка при добавлении расхода. Попробуйте еще раз.")

    except ValueError:
        await update.message.reply_text("Пожалуйста, введите корректное число")
        common.logger.error(f"Ошибка при вводе расхода: {update.message.text.strip()}")

async def backend_add_expense_to_timesheet(context: CallbackContext, expense: float) -> bool:
    """Ваша бэкенд логика для добавления расхода в табель"""
    try:
        sheet = context.user_data.get('sheet')
        selected_category = context.user_data.get('selected_expense_category')
        # Запись данных
        month = context.user_data['selected_date'].strftime("%B")
        row = matches.get_month(month)
        col = matches.get_day_number(selected_category['id'])
        current_value = sheet[f'{col}{row}'].value
        if current_value is None:
            current_value = 0
        if context.user_data['operator'] == "increment":
            sheet[f'{col}{row}'].value = float(current_value) + expense
            logaction = "прибавлено"
            particle = "к"
        elif context.user_data['operator'] == "decrement":
            sheet[f'{col}{row}'].value = float(current_value) - expense
            logaction = "вычтено"
            particle = "из"
        context.user_data.get('workbook').save(common.local_xlsx_path)
        common.logger.info(f"Изменен расход: Категория {context.user_data['selected_expense_category']['name']}, "
            f"Дата: {context.user_data['selected_date']}, {particle} сумме: {current_value} {logaction} {expense}. Результат: {sheet[f'{col}{row}'].value}")
        return True
    except Exception as e:
        common.logger.error(f"Ошибка при изменении расхода: {e}")
        return False