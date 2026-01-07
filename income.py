import common
import os
import matches
from openpyxl import load_workbook
from telegram import Update,InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackContext, CallbackQueryHandler
from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
from datetime import datetime, date, timedelta

async def income(update: Update, context: CallbackContext):
    context.user_data["mode"] = "income"
    keyboard = [
        [InlineKeyboardButton("‚úèÔ∏è –ò–∑–º–µ–Ω–∏—Ç—å –∫–∞—Ç–µ–≥–æ—Ä–∏—é", callback_data=f"get_income_categories")],
        [InlineKeyboardButton("üìÖ –ò–∑–º–µ–Ω–∏—Ç—å –¥–∞—Ç—É", callback_data=f"set_date")],
        [InlineKeyboardButton("‚ûï –î–æ–±–∞–≤–∏—Ç—å –¥–æ—Ö–æ–¥", callback_data=f"ask_for_income_increment")],
        [InlineKeyboardButton("‚ûñ –£–º–µ–Ω—å—à–∏—Ç—å –¥–æ—Ö–æ–¥", callback_data=f"ask_for_income_decrement")],
        [InlineKeyboardButton("üîç –ü–æ—Å–º–æ—Ç—Ä–µ—Ç—å –¥–æ—Ö–æ–¥", callback_data=f"get_income_view")]
    ]

    if not os.path.isfile(common.local_xlsx_path):
        client = common.get_ya_client(context._user_id)
        with client:
            client.download(common.ya_xlsx_path, common.local_xlsx_path)

    workbook = load_workbook(common.local_xlsx_path)
    sheet = workbook[common.year]
    context.user_data['sheet'] = sheet
    context.user_data['workbook'] = workbook

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        if context.user_data.get("selected_income_category") is None:
            context.user_data['selected_income_category'] = common.DEFAULT_INCOME_INFO
        if context.user_data.get("selected_date") is None:
            context.user_data['selected_date'] = date.today()
    else:
        context.user_data['selected_income_category'] = common.DEFAULT_INCOME_INFO
        context.user_data['selected_date'] = date.today()

    if update.callback_query:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            "–î–æ—Ö–æ–¥\n"
            f"–ö–∞—Ç–µ–≥–æ—Ä–∏—è: {context.user_data['selected_income_category']['name']}\n"
            f"–î–∞—Ç–∞: {context.user_data['selected_date']}\n\n",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            "–î–æ—Ö–æ–¥\n"
            f"–ö–∞—Ç–µ–≥–æ—Ä–∏—è: {context.user_data['selected_income_category']['name']}\n"
            f"–î–∞—Ç–∞: {context.user_data['selected_date']}\n\n",
            reply_markup=reply_markup
        )

async def get_income_categories(update: Update, context: CallbackContext):
    sheet = context.user_data['sheet']
    category_list = []
    index = 21
    for col in sheet.iter_cols(min_col=21,min_row=2, max_col=24, values_only=True):
        category_list.append({"id": index, "name": col[0], "month": matches.get_month(index)})
        index += 1

    context.user_data['category_list'] = category_list
    
    keyboard = []
   
    for category in category_list:
        keyboard.append([
            InlineKeyboardButton(
                f"{category['name']}",
                callback_data=f"select_income_category_{category['id']}"
            )
        ])
    keyboard.append([InlineKeyboardButton("‚¨ÖÔ∏è –ù–∞–∑–∞–¥", callback_data="back_to_income")])
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            "–í—ã–±–µ—Ä–∏—Ç–µ –∫–∞—Ç–µ–≥–æ—Ä–∏—é:",
            reply_markup=reply_markup
        )

async def select_income_category_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    await query.answer()

    # –ò–∑–≤–ª–µ–∫–∞–µ–º ID –∫–∞—Ç–µ–≥–æ—Ä–∏–∏ –∏–∑ callback_data
    category_id = int(query.data.replace("select_income_category_", ""))

    # –ù–∞—Ö–æ–¥–∏–º –∫–∞—Ç–µ–≥–æ—Ä–∏—é –≤ —Å–ø–∏—Å–∫–µ
    category_list = context.user_data.get('category_list', [])
    selected_category = next((c for c in category_list if c['id'] == category_id), None)

    if not selected_category:
        await query.edit_message_text("–ö–∞—Ç–µ–≥–æ—Ä–∏—è –Ω–µ –Ω–∞–π–¥–µ–Ω–∞!")
        return

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤—ã–±—Ä–∞–Ω–Ω—É—é –∫–∞—Ç–µ–≥–æ—Ä–∏—é –≤ –∫–æ–Ω—Ç–µ–∫—Å—Ç–µ
    context.user_data['selected_income_category'] = selected_category
    await income(update, context)

async def process_calendar_callback(update: Update, context: CallbackContext):
    """–û–±—Ä–∞–±–æ—Ç—á–∏–∫ –≤—ã–±–æ—Ä–∞ –¥–∞—Ç—ã –∏–∑ –∫–∞–ª–µ–Ω–¥–∞—Ä—è"""
    query = update.callback_query
    await query.answer()
    
    calendar = context.user_data.get('calendar')
    
    if not calendar:
        calendar = DetailedTelegramCalendar()
    
    # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –≤—ã–±–æ—Ä –≤ –∫–∞–ª–µ–Ω–¥–∞—Ä–µ
    result, key, step = calendar.process(query.data)
    
    if not result and key:
        # –û–±–Ω–æ–≤–ª—è–µ–º –∫–∞–ª–µ–Ω–¥–∞—Ä—å –¥–ª—è —Å–ª–µ–¥—É—é—â–µ–≥–æ —à–∞–≥–∞
        await query.edit_message_text(
            f"üìÖ –í—ã–±–µ—Ä–∏—Ç–µ –¥–∞—Ç—É ({LSTEP[step]}):",
            reply_markup=key
        )
    elif result:
        # –î–∞—Ç–∞ –≤—ã–±—Ä–∞–Ω–∞
        selected_date = result
        context.user_data['selected_date'] = selected_date
        
        await income(update, context)

async def ask_for_income(update: Update, context: CallbackContext):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ–º –¥–æ—Ö–æ–¥"""
    
    query = update.callback_query
    await query.answer()

    operator = query.data.replace("ask_for_income_", "")
    context.user_data['operator'] = operator

    keyboard = [
        [InlineKeyboardButton("‚¨ÖÔ∏è –ù–∞–∑–∞–¥", callback_data="back_to_income")]
    ]
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    await query.edit_message_text(
        f"–ö–∞—Ç–µ–≥–æ—Ä–∏—è: {context.user_data['selected_income_category']['name']}\n"
        f"–î–∞—Ç–∞: {context.user_data['selected_date']}\n\n"
        "–í–≤–µ–¥–∏—Ç–µ —Å—É–º–º—É –¥–æ—Ö–æ–¥–∞:",
        reply_markup=reply_markup)

async def get_income_view(update: Update, context: CallbackContext):
    """–ó–∞–ø—Ä–∞—à–∏–≤–∞–µ–º –ø—Ä–æ—Å–º–æ—Ç—Ä –¥–æ—Ö–æ–¥–∞"""
    query = update.callback_query
    await query.answer()

    keyboard = [
        [InlineKeyboardButton("‚¨ÖÔ∏è –ù–∞–∑–∞–¥", callback_data="back_to_income")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        sheet = context.user_data.get('sheet')
        selected_category = context.user_data.get('selected_income_category')
        # –ó–∞–ø–∏—Å—å –¥–∞–Ω–Ω—ã—Ö
        month = context.user_data['selected_date'].strftime("%B")
        row = matches.get_month(month)
        col = matches.get_letter(selected_category['id'])
        current_value = sheet[f'{col}{row}'].value
        if current_value is None:
            current_value = 0
        
        await query.edit_message_text(
                f"‚úÖ –£—Å–ø–µ—à–Ω–æ!\n\n"
                f"–ö–∞—Ç–µ–≥–æ—Ä–∏—è: {context.user_data['selected_income_category']['name']}\n"
                f"üìÖ –î–∞—Ç–∞: {context.user_data['selected_date']}\n"
                f"üí∞ –¢–µ–∫—É—â–∏–π –¥–æ—Ö–æ–¥: {current_value}\n\n"
                "–ß—Ç–æ –¥–∞–ª—å—à–µ?",
                reply_markup=reply_markup
            )

    except ValueError:
        await update.message.reply_text("–ü—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –∏–Ω—Ñ—ã –æ –¥–æ—Ö–æ–¥–µ –ø—Ä–æ–∏–∑–æ—à–ª–∞ –æ—à–∏–±–∫–∞.")


async def process_income_input(update: Update, context: CallbackContext):
    """–û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ—Ç –≤–≤–µ–¥–µ–Ω–Ω–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —á–∞—Å–æ–≤"""

    try:
        # –ü–∞—Ä—Å–∏–º –≤–≤–µ–¥–µ–Ω–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
        income_text = update.message.text.strip()
        income = float(update.message.text.strip().replace(',', '.'))
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤–∞–ª–∏–¥–Ω–æ—Å—Ç—å
        if income <= 0:
            await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ –ø–æ–ª–æ–∂–∏—Ç–µ–ª—å–Ω–æ–µ —á–∏—Å–ª–æ:")
            return
        
        success = await backend_add_income_to_timesheet(context=context, income=income)
        
        if success:
            keyboard = [
                [InlineKeyboardButton("‚ûï –î–æ–±–∞–≤–∏—Ç—å –¥–æ—Ö–æ–¥", callback_data=f"ask_for_income_increment")],
                [InlineKeyboardButton("‚ûñ –£–º–µ–Ω—å—à–∏—Ç—å –¥–æ—Ö–æ–¥", callback_data=f"ask_for_income_decrement")],
                [InlineKeyboardButton("‚¨ÖÔ∏è –ù–∞–∑–∞–¥", callback_data="back_to_income")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                f"‚úÖ –£—Å–ø–µ—à–Ω–æ –¥–æ–±–∞–≤–ª–µ–Ω–æ!\n\n"
                f"–ö–∞—Ç–µ–≥–æ—Ä–∏—è: {context.user_data['selected_income_category']['name']}\n"
                f"üìÖ –î–∞—Ç–∞: {context.user_data['selected_date']}\n"
                f"üí∞ –î–æ—Ö–æ–¥: {income}\n\n"
                "–ß—Ç–æ –¥–∞–ª—å—à–µ?",
                reply_markup=reply_markup
            )
            client = common.get_ya_client(context._user_id)
            with client:
                client.upload(common.local_xlsx_path, common.ya_xlsx_path,overwrite=True)
            os.remove(common.local_xlsx_path)

        else:
            await update.message.reply_text("‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–æ–±–∞–≤–ª–µ–Ω–∏–∏ –¥–æ—Ö–æ–¥–∞. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –µ—â–µ —Ä–∞–∑.")

    except ValueError:
        await update.message.reply_text("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤–≤–µ–¥–∏—Ç–µ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ–µ —á–∏—Å–ª–æ")

async def backend_add_income_to_timesheet(context: CallbackContext, income: float) -> bool:
    """–í–∞—à–∞ –±—ç–∫–µ–Ω–¥ –ª–æ–≥–∏–∫–∞ –¥–ª—è –¥–æ–±–∞–≤–ª–µ–Ω–∏—è –¥–æ—Ö–æ–¥–∞ –≤ —Ç–∞–±–µ–ª—å"""
    try:
        sheet = context.user_data.get('sheet')
        selected_category = context.user_data.get('selected_income_category')
        # –ó–∞–ø–∏—Å—å –¥–∞–Ω–Ω—ã—Ö
        month = context.user_data['selected_date'].strftime("%B")
        row = matches.get_month(month)
        col = matches.get_letter(selected_category['id'])
        current_value = sheet[f'{col}{row}'].value
        if current_value is None:
            current_value = 0
        if context.user_data['operator'] == "increment":
            sheet[f'{col}{row}'].value = float(current_value) + income
            logaction = "–ø—Ä–∏–±–∞–≤–ª–µ–Ω–æ"
            particle = "–∫"
        elif context.user_data['operator'] == "decrement":
            sheet[f'{col}{row}'].value = float(current_value) - income
            logaction = "–≤—ã—á—Ç–µ–Ω–æ"
            particle = "–∏–∑"
        context.user_data.get('workbook').save(common.local_xlsx_path)
        common.logger.info(f"–ò–∑–º–µ–Ω–µ–Ω –¥–æ—Ö–æ–¥: –ö–∞—Ç–µ–≥–æ—Ä–∏—è {context.user_data['selected_income_category']['name']}, "
            f"–î–∞—Ç–∞: {context.user_data['selected_date']}, {particle} —Å—É–º–º–µ: {current_value} {logaction} {income}. –†–µ–∑—É–ª—å—Ç–∞—Ç: {sheet[f'{col}{row}'].value}")
        return True
    except Exception as e:
        common.logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∏–∑–º–µ–Ω–µ–Ω–∏–∏ –¥–æ—Ö–æ–¥–∞: {e}")
        return False