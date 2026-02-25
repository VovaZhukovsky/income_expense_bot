import common
import matches
import os
from openpyxl import load_workbook
from telegram import Update,InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import CallbackContext
from telegram_bot_calendar import DetailedTelegramCalendar, LSTEP
from datetime import date

async def action(update: Update, context: CallbackContext, mode: common.Mode = common.Mode.NONE):

    if update.callback_query:
        mode = context.user_data["mode"]
    else:
        context.user_data["mode"] = mode

    match mode.value:
        case "income":
            context.user_data["action_info"] = common.DEFAULT_INCOME_INFO
        case "expense":
            context.user_data["action_info"] = common.DEFAULT_EXPENSE_INFO

    keyboard = [
        [InlineKeyboardButton("✏️ Change category", callback_data="get_categories")],
        [InlineKeyboardButton("📅 Change datetime", callback_data="set_date")],
        [InlineKeyboardButton(f"➕ Increase {mode.value}", callback_data="ask_for_increment")],
        [InlineKeyboardButton(f"➖ Decrease {mode.value}", callback_data="ask_for_decrement")],
        [InlineKeyboardButton(f"🔍 Show {mode.value}", callback_data="get_view")]
    ]
    if not os.path.isfile(common.local_xlsx_path):
        client = common.get_ya_client(context._user_id)
        with client:
            client.download(common.ya_xlsx_path, common.local_xlsx_path)
    workbook = load_workbook(common.local_xlsx_path)
    sheet = workbook[common.year]
    context.user_data['sheet'] = sheet
    context.user_data['workbook'] = workbook

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        if context.user_data.get("selected_category") is None:
            context.user_data['selected_category'] = context.user_data["action_info"]
        if context.user_data.get("selected_date") is None:
            context.user_data['selected_date'] = date.today()
    else:
        context.user_data['selected_category'] = context.user_data["action_info"]
        context.user_data['selected_date'] = date.today()

    if update.callback_query:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            f"{mode.value.capitalize()}\n"
            f"Category: {context.user_data['selected_category']['name']}\n"
            f"Datetime: {context.user_data['selected_date']}\n\n",
            reply_markup=reply_markup
        )
    else:
        await update.message.reply_text(
            f"{mode.value.capitalize()}\n"
            f"Category: {context.user_data['selected_category']['name']}\n"
            f"Datetime: {context.user_data['selected_date']}\n\n",
            reply_markup=reply_markup
        )


async def get_categories(update: Update, context: CallbackContext):
    sheet = context.user_data['sheet']
    category_list = []
    action_info = context.user_data['action_info']
    index = action_info["min_col"]
    for col in sheet.iter_cols(min_col=action_info["min_col"], min_row=action_info["min_row"], max_col=action_info["max_col"], values_only=True):
        category_list.append({"id": index, "name": col[0], "month": matches.get_month(index)})
        index += 1

    context.user_data['category_list'] = category_list

    keyboard = []

    for category in category_list:
        keyboard.append([
            InlineKeyboardButton(
                f"{category['name']}",
                callback_data=f"select_category_{category['id']}"
            )
        ])
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="back")])

    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.callback_query:
        query = update.callback_query
        await query.answer()
        await query.edit_message_text(
            "Choose category:",
            reply_markup=reply_markup
        )

async def select_category_handler(update: Update, context: CallbackContext):
    query = update.callback_query
    await query.answer()

    # Извлекаем ID категории из callback_data
    category_id = int(query.data.replace("select_category_", ""))

    # Находим категорию в списке
    category_list = context.user_data.get('category_list', [])
    selected_category = next((c for c in category_list if c['id'] == category_id), None)

    if not selected_category:
        await query.edit_message_text("Category not found!")
        return

    # Сохраняем выбранную категорию в контексте
    context.user_data['selected_category'] = selected_category
    await action(update, context,context.user_data["mode"])


async def process_calendar_callback(update: Update, context: CallbackContext):
    """Обработчик выбора даты из календаря"""
    query = update.callback_query
    await query.answer()

    calendar = context.user_data.get('calendar')

    if not calendar:
        calendar = DetailedTelegramCalendar()

    # Обрабатываем выбор в календаре
    result, key, step = calendar.process(query.data)

    if not result and key:
        # Обновляем календарь для следующего шага
        await query.edit_message_text(
            f"📅 Choose datetime ({LSTEP[step]}):",
            reply_markup=key
        )
    elif result:
        # Дата выбрана
        selected_date = result
        context.user_data['selected_date'] = selected_date

        await action(update, context,context.user_data["mode"])


async def ask_for(update: Update, context: CallbackContext):

    query = update.callback_query
    await query.answer()

    operator = query.data.replace("ask_for_", "")
    context.user_data['operator'] = operator

    keyboard = [
        [InlineKeyboardButton("⬅️ Back", callback_data="back")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        f"Category: {context.user_data['selected_category']['name']}\n"
        f"Datetime: {context.user_data['selected_date']}\n\n"
        f"Enter sum {context.user_data["mode"].value}:",
        reply_markup=reply_markup)


async def get_view(update: Update, context: CallbackContext):
    """Запрашиваем просмотр дохода"""
    query = update.callback_query
    await query.answer()

    keyboard = [
        [InlineKeyboardButton("⬅️ Back", callback_data="back")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        sheet = context.user_data.get('sheet')
        selected_category = context.user_data.get('selected_category')
        # Запись данных
        month = context.user_data['selected_date'].strftime("%B")
        row = matches.get_month(month)
        col = matches.get_day_number(selected_category['id'])
        current_value = sheet[f'{col}{row}'].value
        if current_value is None:
            current_value = 0

        await query.edit_message_text(
            f"✅ Success!\n\n"
            f"Category: {context.user_data['selected_category']['name']}\n"
            f"📅 Datetime: {context.user_data['selected_date']}\n"
            f"💰 Current {context.user_data["mode"].value}: {current_value}\n\n"
            "What is next?",
            reply_markup=reply_markup
        )

    except ValueError:
        await update.message.reply_text("An error occurred while getting information.")


async def process_input(update: Update, context: CallbackContext):
    """Обрабатывает введенное количество часов"""

    try:
        # Парсим введенное значение
        value = float(update.message.text.strip().replace(',', '.'))
        mode = context.user_data["mode"]
        # Проверяем валидность
        if value <= 0:
            await update.message.reply_text("Please, enter positive number:")
            return

        backend_result = await backend_timesheet(context=context, value=value)

        if backend_result.result:
            keyboard = [
                [InlineKeyboardButton(f"➕ increase {mode.value}", callback_data=f"ask_for_increment")],
                [InlineKeyboardButton(f"➖ decrease {mode.value}", callback_data=f"ask_for_decrement")],
                [InlineKeyboardButton("⬅️ back", callback_data="back")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)

            await update.message.reply_text(
                f"✅ Success!\n\n"
                f"Category: {context.user_data['selected_category']['name']}\n"
                f"📅 Datetime: {context.user_data['selected_date']}\n"
                f"💰 Mode {mode.value}:\n"
                f" Old value: {backend_result.old_value}\n"
                f" New value: {backend_result.new_value}\n"
                f" Diff: {backend_result.diff_value}\n\n"
                "What is next?",
                reply_markup=reply_markup
            )
            client = common.get_ya_client(context._user_id)
            with client:
                client.upload(common.local_xlsx_path, common.ya_xlsx_path, overwrite=True)
            os.remove(common.local_xlsx_path)

        else:
            await update.message.reply_text(f"❌ An error occurred while interacting with {mode.value}. Try again.")

    except ValueError:
        await update.message.reply_text("Please, enter positive number!")

async def backend_timesheet(context: CallbackContext, value: float) -> common.Backend_TimeShift_Result:
    """Ваша бэкенд логика для добавления дохода в табель"""

    try:
        log_action = ""
        mode = context.user_data["mode"]
        sheet = context.user_data.get('sheet')
        selected_category = context.user_data.get('selected_category')
        # Запись данных
        month = context.user_data['selected_date'].strftime("%B")
        row = matches.get_month(month)
        col = matches.get_day_number(selected_category['id'])
        current_value = sheet[f'{col}{row}'].value
        if current_value is None:
            current_value = 0
        if context.user_data['operator'] == "increment":
            sheet[f'{col}{row}'].value = float(current_value) + value
            log_action = "increased"
        elif context.user_data['operator'] == "decrement":
            sheet[f'{col}{row}'].value = float(current_value) - value
            log_action = "decreased"
        context.user_data.get('workbook').save(common.local_xlsx_path)
        common.logger.info(f"{mode.value} changed: category {context.user_data['selected_category']['name']}, "
            f"datetime: {context.user_data['selected_date']}: {current_value} {log_action} {value}. result: {sheet[f'{col}{row}'].value}")
        context.user_data['operator'] = None
        return common.Backend_TimeShift_Result(True, current_value, sheet[f'{col}{row}'].value, value)
    except Exception as e:
        common.logger.error(f"An error occurred while interacting with {context.user_data["mode"]}: {e}")
        return common.Backend_TimeShift_Result(False)